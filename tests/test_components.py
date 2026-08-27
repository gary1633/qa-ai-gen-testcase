import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import os
import openpyxl
from src.core.models import (
    RequirementAnalysis,
    AcceptanceCriterion,
    ProductRisk,
    TestScenario,
    TestCase,
    ReviewResult,
    ReviewIssue,
    TraceabilityItem
)
from src.core.linter import lint_test_case, lint_test_suite_coverage, lint_scenarios
from src.utils.file_parsers import extract_input_content
from src.utils.excel_exporter import export_test_cases_to_excel
from src.core.workflow import build_qa_agentic_graph
from src.core.llm import load_qa_rules, load_config
from src.core.prompt_loader import resolve_domain_pack, load_domain_pack


def test_file_parser():
    print("[1/4] Testing File Parsers...")
    content, file_type = extract_input_content("samples/sample_user_story.md")
    assert file_type == "md"
    assert "VWCBT-4102" in content
    assert "AC-01" in content

    # Test Multi-Source with User Clarifications / Extra Info
    from src.utils.file_parsers import merge_multiple_sources
    merged_text, f_type, meta = merge_multiple_sources([
        "samples/sample_user_story.md",
        "Lưu ý: Khung giờ EOD chuẩn là 18h VNT, phí cố định 5.000 VND"
    ])
    assert f_type == "multi_document"
    assert "THÔNG TIN BỔ SUNG / LÀM RÕ TỪ USER" in merged_text or "User Clarifications & Overrides" in merged_text
    assert "18h VNT" in merged_text
    print("  -> Parser OK! Multi-source & User Clarifications tested (Read", len(merged_text), "chars).")

def test_linter():
    print("\n[2/4] Testing Deterministic QA Linter...")
    bad_tc = TestCase(
        testcase_id="TC 99",
        group_feature="1. Feature",
        group_functional="1.1 Functional",
        title="Bad test",
        preconditions="None",
        steps="Click login and wait a bit",
        expected_result="Verify it works properly",
        actual_result="",
        test_data="valid data",
        creator="QA",
        test_date="24/08/2026",
        test_status="Not Test",
        priority="High",
        plan_execution="Sprint",
        executed_date="",
        note="AC-01"
    )
    issues = lint_test_case(bad_tc)
    issue_types = [i.issue_type for i in issues]
    assert "Non-Deterministic Expected Result" in issue_types or "Ambiguous Step" in issue_types
    assert any("wait a bit" in i.description for i in issues)
    assert any("verify it works" in i.description for i in issues)
    print(f"  -> Linter correctly flagged {len(issues)} issues in ambiguous testcase!")

    # Test Banking Domain Linter checks (Missing Idempotency & Conditional QĐ 2345)
    payment_analysis = RequirementAnalysis(
        feature_name="Chuyển tiền Napas 24/7",
        business_overview="Thanh toán chuyển tiền",
        banking_domain="Payments & Fund Transfers (Napas, VietQR, Swift)",
        acceptance_criteria=[
            AcceptanceCriterion(
                ac_id="AC-01",
                title="Chuyển tiền qua App",
                description="Chuyển tiền trên App yêu cầu xác thực sinh trắc học khuôn mặt",
                risk_level="High"
            )
        ],
        product_risks=[
            ProductRisk(
                risk_id="RSK-01",
                risk_title="Trừ tiền 2 lần do timeout",
                risk_category="Integration & Timeout Risk",
                likelihood=4, impact=5, risk_score=20,
                risk_level="Critical", linked_ac_id="AC-01",
                mitigation_test_focus="Kiểm thử Idempotency key và timeout."
            )
        ]
    )
    incomplete_tcs = [
        TestCase(
            testcase_id="TC 01", group_feature="1. Transfer", group_functional="1.1 Basic",
            title="Chuyển tiền đơn giản 100k", preconditions="Active",
            steps="1. Gửi request POST /transfer\n2. Nhận kết quả",
            expected_result="HTTP Status 200 OK thành công.",
            actual_result="", test_data='{"amount": 100000}',
            creator="QA", test_date="24/08/2026", test_status="Not Test",
            priority="High", plan_execution="Sprint", executed_date="", note="AC-01"
        )
    ]
    coverage_issues = lint_test_suite_coverage(payment_analysis, incomplete_tcs)
    cov_issue_types = [i.issue_type for i in coverage_issues]
    assert "Missing Idempotency/Timeout Case" in cov_issue_types
    assert "Banking Compliance Violation (QĐ 2345)" in cov_issue_types
    print(f"  -> Banking Linter correctly flagged {len(coverage_issues)} banking domain issues (Idempotency & Conditional QĐ 2345)!")
    # Test Traceability Matrix Model
    trace_item = TraceabilityItem(
        ac_id="AC-01",
        ac_title="Chuyển tiền Napas",
        risk_level="High",
        covered_test_cases=["TC 01", "TC 02"],
        coverage_status="COVERED",
        coverage_notes="Covered Positive & Negative Boundary"
    )
    assert trace_item.ac_id == "AC-01"
    assert trace_item.coverage_status == "COVERED"
    print("  -> Traceability Matrix Data Model verified!")

def test_excel_exporter():
    print("\n[3/4] Testing Excel Exporter with EF_TestCases.xlsx (Standard Template with Logo)...")
    req_analysis = RequirementAnalysis(
        feature_name="VWCBT-4102: Chuyển tiền Napas 24/7",
        app_name="Branch Portal / Mobile Banking",
        version="UAT 3.1.0",
        jira_or_doc_link="https://galaxyfinx.atlassian.net/browse/VWCBT-4102",
        business_overview="Chuyển tiền nhanh liên ngân hàng 24/7",
        acceptance_criteria=[
            AcceptanceCriterion(
                ac_id="AC-01",
                title="Chuyển tiền qua STK",
                description="Chuyển tiền qua STK thành công",
                business_rules=["Tối thiểu 10k, tối đa 500tr"],
                risk_level="High"
            )
        ]
    )

    sample_test_cases = [
        TestCase(
            testcase_id="TC 01",
            group_feature="1. Chuyển tiền Napas 247 qua Số tài khoản (AC-01)",
            group_functional="1.1. Luồng thành công (Happy Path & Phí giao dịch)",
            title="Chuyển tiền thành công dưới 1 triệu VND (Miễn phí giao dịch)",
            preconditions="Tài khoản nguồn Active, số dư khả dụng >= 500,000 VND. Ngân hàng đích liên kết Napas.",
            steps="1. Gửi request POST /v1/transfer/napas247\n2. Kiểm tra status code và response body",
            expected_result="- HTTP Status: 200 OK\n- Response Body:\n{\n  \"status\": \"SUCCESS\",\n  \"fee\": 0,\n  \"trace_no\": \"NP20260824001\"\n}",
            actual_result="",
            test_data="{\n  \"from_account\": \"1012345678\",\n  \"to_account\": \"9988776655\",\n  \"to_bank_code\": \"VCB\",\n  \"amount\": 500000\n}",
            creator="QA Agent",
            test_date="24/08/2026",
            test_status="Not Test",
            priority="High",
            plan_execution="Sprint 1",
            executed_date="",
            note="AC-01"
        ),
        TestCase(
            testcase_id="TC 02",
            group_feature="1. Chuyển tiền Napas 247 qua Số tài khoản (AC-01)",
            group_functional="1.1. Luồng thành công (Happy Path & Phí giao dịch)",
            title="Chuyển tiền thành công từ 1 triệu VND trở lên (Thu phí 2,200 VND)",
            preconditions="Tài khoản nguồn Active, số dư khả dụng >= 2,002,200 VND.",
            steps="1. Gửi request POST /v1/transfer/napas247\n2. Kiểm tra trừ tiền gốc và phí",
            expected_result="- HTTP Status: 200 OK\n- Phí giao dịch: 2,200 VND\n- Số dư tài khoản nguồn bị trừ: 2,002,200 VND",
            actual_result="",
            test_data="{\n  \"from_account\": \"1012345678\",\n  \"amount\": 2000000\n}",
            creator="QA Agent",
            test_date="24/08/2026",
            test_status="Not Test",
            priority="High",
            plan_execution="Sprint 1",
            executed_date="",
            note="AC-01"
        ),
        TestCase(
            testcase_id="TC 03",
            group_feature="2. Kiểm tra Hạn mức Giao dịch (AC-02)",
            group_functional="2.1. Phân tích giá trị biên (Boundary Value Analysis)",
            title="Kiểm tra chuyển số tiền dưới mức tối thiểu (9,999 VND)",
            preconditions="Tài khoản nguồn hợp lệ.",
            steps="1. Gửi request POST /v1/transfer/napas247 với amount = 9999\n2. Kiểm tra mã lỗi trả về",
            expected_result="- HTTP Status: 400 Bad Request\n- Error Code: ERR_MIN_AMOUNT\n- Error Message: 'Số tiền chuyển tối thiểu là 10,000 VND'",
            actual_result="",
            test_data="{\n  \"amount\": 9999\n}",
            creator="QA Agent",
            test_date="24/08/2026",
            test_status="Not Test",
            priority="Medium",
            plan_execution="Sprint 1",
            executed_date="",
            note="AC-02"
        )
    ]

    out_file = export_test_cases_to_excel(
        analysis=req_analysis,
        test_cases=sample_test_cases,
        template_path="EF_TestCases.xlsx",
        output_path=None,  # Tự động xuất ra outputs/
        target_sheet_name="VWCBT_4102_Napas_Test"
    )

    # Verify generated sheet
    wb = openpyxl.load_workbook(out_file, data_only=False)
    print("  -> Exported file sheet names:", wb.sheetnames)
    assert "Workstream Progress Summary" in wb.sheetnames
    assert "VWCBT_4102_Napas_Test" in wb.sheetnames
    assert "Tài liệu tổng hợp" in wb.sheetnames
    # Template phiếu KTNV or standard common sheets
    assert "Daily Execution Log" in wb.sheetnames
    assert "Checklist Report" in wb.sheetnames
    assert "RAW JIRA Bug" in wb.sheetnames
    assert "EF_BLOCKADE_SA" not in wb.sheetnames
    assert "EF_Calculate_Savings_Interest" not in wb.sheetnames
    
    ws = wb["VWCBT_4102_Napas_Test"]
    # Check metadata
    assert ws.cell(10, 3).value == "Branch Portal / Mobile Banking"
    assert ws.cell(11, 3).value == "UAT 3.1.0"
    
    # Check formula in row 19
    total_formula = ws.cell(19, 1).value
    print("  -> Generated formula for Total:", total_formula)
    assert total_formula.startswith("=COUNTIF(A22:A")
    
    print(f"  -> Excel Exporter OK! Sheet 'VWCBT_4102_Napas_Test' created and populated.")


def test_graph_compilation():
    print("\n[4/4] Testing LangGraph Workflow Compilation...")
    graph = build_qa_agentic_graph()
    assert graph is not None
    print("  -> LangGraph StateGraph compiled successfully!")
    # Test RequirementAnalysis attribute compatibility
    req = RequirementAnalysis(
        feature_name="Test Feature",
        business_overview="Mục tiêu nghiệp vụ kiểm thử",
        banking_domain="Payments & Fund Transfers (Napas, VietQR, Swift)"
    )
    assert req.business_overview == "Mục tiêu nghiệp vụ kiểm thử"
    assert req.business_objective == "Mục tiêu nghiệp vụ kiểm thử"
    print("  -> RequirementAnalysis business_overview/business_objective compatibility verified!")
    # Test ProductRisk model fields
    risk = ProductRisk(
        risk_id="RSK-01",
        risk_title="Rủi ro trừ tiền 2 lần",
        risk_description="Hệ thống có thể gửi duplicate request dẫn đến trừ tiền 2 lần",
        risk_category="Financial & Ledger Risk",
        risk_level="Critical",
        mitigation_test_focus="Test Idempotency 50ms"
    )
    assert risk.risk_title == "Rủi ro trừ tiền 2 lần"
    assert risk.risk_description.startswith("Hệ thống có thể")
    print("  -> ProductRisk risk_title/risk_description verified!")
    # Test cross-property aliases for TestScenario & TestCase
    sc = TestScenario(
        scenario_id="SC_01",
        scenario_title="Kiểm tra chuyển tiền thành công",
        group_feature="Chuyển tiền",
        group_functional="Napas 24/7"
    )
    assert sc.title == "Kiểm tra chuyển tiền thành công"
    assert sc.scenario_title == "Kiểm tra chuyển tiền thành công"

    tc = TestCase(
        testcase_id="TC 01",
        title="Kiểm tra chuyển tiền thành công khi truyền đúng payload",
        steps="1. Gửi request",
        expected_result="Status 200 OK",
        test_data="{}"
    )
    assert tc.title == "Kiểm tra chuyển tiền thành công khi truyền đúng payload"
    assert tc.scenario_title == "Kiểm tra chuyển tiền thành công khi truyền đúng payload"
    print("  -> TestScenario/TestCase title & scenario_title cross-aliases verified!")
    # Test clean_jira_key_from_title
    from src.utils.file_parsers import clean_jira_key_from_title
    dirty_title = 'Kiểm tra thực thi giao dịch VWCBT-3230 thành công khi truyền trường "transaction_mode" mang giá trị hợp lệ "standard"'
    cleaned_title = clean_jira_key_from_title(dirty_title)
    assert "VWCBT-3230" not in cleaned_title
    assert cleaned_title == 'Kiểm tra thực thi giao dịch thành công khi truyền trường "transaction_mode" mang giá trị hợp lệ "standard"'
    print("  -> clean_jira_key_from_title successfully stripped embedded Jira key!")
    # Test Guardrail Validation
    from src.core.guardrail import validate_requirement_input
    assert not validate_requirement_input("alo")[0]
    assert not validate_requirement_input("hi")[0]
    assert not validate_requirement_input("test")[0]
    assert not validate_requirement_input("123")[0]
    assert not validate_requirement_input("asdfghjkl")[0]
    assert not validate_requirement_input("lam test case ho voi")[0]
    
    # Valid cases
    assert validate_requirement_input("VWCBT-3648")[0]
    assert validate_requirement_input("https://galaxyfinx.atlassian.net/browse/VWCBT-3648")[0]
    assert validate_requirement_input("Tính năng chuyển tiền Napas 24/7. AC1: Hạn mức tối thiểu 10,000 VND. AC2: Hạch toán nợ có.")[0]
    print("  -> Input Guardrail & Spam Prevention verified!")

    # Test Clarification Gate Conditional Routing
    from src.core.workflow import should_continue_after_analysis
    clarify_analysis = RequirementAnalysis(
        feature_name="Tính năng chưa rõ",
        needs_user_clarification=True,
        clarification_questions=["Hạn mức tối đa áp dụng theo ngày hay theo từng giao dịch?"]
    )
    assert should_continue_after_analysis({"requirement_analysis": clarify_analysis}) == "needs_clarification"

    clear_analysis = RequirementAnalysis(
        feature_name="Tính năng rõ ràng",
        needs_user_clarification=False
    )
    assert should_continue_after_analysis({"requirement_analysis": clear_analysis}) == "design_scenarios"
    print("  -> Interactive Clarification Gate (Human-in-the-Loop) routing verified!")
def test_agent_invocations():
    print("\n[5/5] Testing Agent LLM Invocation Signatures & Contracts...")
    from unittest.mock import patch
    from src.agents.requirement_analyst import analyze_requirements
    from src.agents.scenario_designer import design_test_scenarios, ScenarioListResponse
    from src.agents.testcase_generator import generate_test_cases, BatchTestSuiteResponse
    from src.agents.reviewer import review_and_lint_test_suite, SemanticReviewPayload
    with patch("src.agents.requirement_analyst.invoke_structured_llm") as mock_ana, \
         patch("src.agents.scenario_designer.invoke_structured_llm") as mock_sc, \
         patch("src.agents.testcase_generator.invoke_structured_llm") as mock_gen, \
         patch("src.agents.reviewer.invoke_structured_llm") as mock_rev:
        mock_ana.return_value = RequirementAnalysis(
            feature_name="Chặn rút tiền EOD",
            banking_domain="Savings & Term Deposits (Interest, Maturity, Accrual)",
            acceptance_criteria=[AcceptanceCriterion(ac_id="AC-01", title="Chặn EOD", description="Chặn lúc 18h", risk_level="High")]
        )
        analysis = analyze_requirements("Chặn rút tiền trong giờ EOD 18:00 VNT", provider="gemini")
        assert mock_ana.call_count == 1
        assert "system_prompt" in mock_ana.call_args.kwargs
        assert "user_prompt" in mock_ana.call_args.kwargs
        assert "schema" in mock_ana.call_args.kwargs

        mock_sc.return_value = ScenarioListResponse(
            scenarios=[
                TestScenario(scenario_id="SC_01", trace_ac_id="AC-01", group_feature="1. Chặn EOD", group_functional="1.1. Luồng chính", scenario_title='Kiểm tra "18:00:00"', testing_technique="Boundary Value Analysis (BVA)", priority="High")
            ]
        )
        scenarios = design_test_scenarios(analysis, provider="gemini")
        assert mock_sc.call_count == 1
        assert "system_prompt" in mock_sc.call_args.kwargs

        mock_gen.return_value = BatchTestSuiteResponse(
            test_cases=[
                TestCase(testcase_id="TC 01", group_feature="1. Chặn EOD", group_functional="1.1. Luồng chính", title='Kiểm tra "18:00:00"', preconditions="Active", steps="POST", expected_result="400 CV_043", actual_result="", test_data="{}", creator="QA", test_date="24/08/2026", test_status="Not Test", priority="High", plan_execution="Release", executed_date="", note="AC-01")
            ]
        )
        gen_result = generate_test_cases(analysis, scenarios, provider="gemini")
        assert mock_gen.call_count == 1
        assert "system_prompt" in mock_gen.call_args.kwargs
        test_cases = gen_result.test_cases
        assert len(test_cases) == 1

        mock_rev.return_value = SemanticReviewPayload(
            semantic_score=100,
            traceability_matrix=[],
            semantic_issues=[],
            feedback_summary="All good"
        )
        rev_res = review_and_lint_test_suite(analysis, test_cases, provider="gemini")
        assert mock_rev.call_count == 1
        assert "system_prompt" in mock_rev.call_args.kwargs
        assert isinstance(rev_res, ReviewResult)
    print("  -> All 4 Agent invoke_structured_llm signatures & kwargs verified!")
def test_multi_domain_support():
    print("\n[6/6] Testing Universal Multi-Domain Support (E-Commerce, Logistics, SaaS, FinTech)...")
    from src.core.models import RequirementAnalysis, AcceptanceCriterion, TestCase
    from src.core.linter import lint_test_suite_coverage

    # 1. Test E-Commerce Domain
    ecom_analysis = RequirementAnalysis(
        feature_name="Áp dụng Voucher giảm giá Giỏ hàng & Flash Sale",
        app_name="ShopeeClone Web / Mobile",
        version="2.0.0",
        banking_domain="E-Commerce & Retail (Cart, Checkout, Promotion, Flash Sale)",
        business_overview="Khách hàng áp mã giảm giá 20% tối đa 50k khi giỏ hàng >= 200k",
        acceptance_criteria=[
            AcceptanceCriterion(ac_id="AC-01", title="Áp voucher hợp lệ", description="Giảm 20% khi đơn >= 200k", risk_level="High"),
            AcceptanceCriterion(ac_id="AC-02", title="Chặn voucher khi đơn < 200k", description="Báo lỗi khi đơn không đủ điều kiện", risk_level="Medium")
        ]
    )
    assert ecom_analysis.business_domain == "E-Commerce & Retail (Cart, Checkout, Promotion, Flash Sale)"
    
    ecom_cases = [
        TestCase(testcase_id="TC 01", group_feature="1. Áp voucher (AC-01)", group_functional="1.1. Thành công", title='Kiểm tra áp voucher thành công khi đơn "250000" VND', preconditions="User có voucher", steps="1. Gửi request áp voucher\n2. Kiểm tra giảm giá", expected_result="Giảm 50,000 VND, HTTP 200 OK", test_data='{"cart_total": 250000}', note="AC-01"),
        TestCase(testcase_id="TC 02", group_feature="1. Áp voucher (AC-02)", group_functional="1.2. Bắt lỗi", title='Kiểm tra báo lỗi khi đơn "150000" dưới mức tối thiểu', preconditions="User có voucher", steps="1. Gửi request áp voucher\n2. Kiểm tra lỗi", expected_result="HTTP 400 Bad Request, mã lỗi 'VOUCHER_MIN_ORDER_NOT_MET'", test_data='{"cart_total": 150000}', note="AC-02")
    ]
    ecom_issues = lint_test_suite_coverage(ecom_analysis, ecom_cases)
    assert not any(i.issue_type == "Traceability Gap" for i in ecom_issues)
    print("  -> E-Commerce Domain analysis & testcase coverage verified!")

    # 2. Test Logistics & Supply Chain Domain
    logistics_analysis = RequirementAnalysis(
        feature_name="Cập nhật trạng thái lộ trình vận chuyển đơn hàng",
        app_name="GHTK / ViettelPost Delivery Hub",
        banking_domain="Logistics, Supply Chain & Fleet Tracking",
        business_overview="Tài xế quét mã barcode để chuyển trạng thái từ IN_TRANSIT sang DELIVERED",
        acceptance_criteria=[
            AcceptanceCriterion(ac_id="AC-01", title="Quét giao thành công", description="Cập nhật DELIVERED", risk_level="High")
        ]
    )
    assert logistics_analysis.business_domain == "Logistics, Supply Chain & Fleet Tracking"
    print("  -> Logistics Domain verified!")

def test_linter_dead_checks_regression():
    """
    D2 regression: bộ test suite Healthcare có 1 rủi ro Critical (RSK-01) chưa được bao phủ
    và chỉ có duy nhất 1 test case happy-path (mặc định trả về "HTTP Status 200 OK").
    Trước khi sửa Step 4-5, Linter chỉ phát hiện được 'Technique Under-Coverage (Negative EP)'
    vì BVA-check và RBT-check đều là dead code (luôn no-op) và Healthcare không có domain rule riêng.
    """
    print("\n[7/7] Testing Linter Dead-Checks Regression (BVA / RBT / Domain PHI)...")
    healthcare_analysis = RequirementAnalysis(
        feature_name="Tra cứu hồ sơ bệnh án điện tử",
        banking_domain="Healthcare",
        acceptance_criteria=[
            AcceptanceCriterion(ac_id="AC-01", title="Xem hồ sơ bệnh án", description="Bác sĩ xem được hồ sơ bệnh án của bệnh nhân", risk_level="Critical")
        ],
        product_risks=[
            ProductRisk(risk_id="RSK-01", risk_title="Truy cập trái phép hồ sơ bệnh án", risk_level="Critical", linked_ac_id="AC-01")
        ]
    )
    happy_path_case = TestCase()  # Mặc định: title="Test case", expected_result="HTTP Status 200 OK, xử lý thành công."
    issues = lint_test_suite_coverage(healthcare_analysis, [happy_path_case])
    issue_types = [i.issue_type for i in issues]

    assert "Technique Under-Coverage (Negative EP)" in issue_types
    assert "Technique Under-Coverage (BVA)" in issue_types
    assert "RBT Under-Coverage Violation" in issue_types
    assert "Missing PHI Access-Control Case" in issue_types
    print(f"  -> Linter correctly flagged all {len(issue_types)} previously-dead checks: {issue_types}")


def test_new_qa_capabilities():
    """
    Kiểm chứng trực tiếp 6 năng lực mới của bộ khung Linter/Config đã bổ sung:
    Assertion-Anchor, Duplicate Detection, Scenario-Level Lint, Live qa_rules, Domain Pack Routing,
    Phantom AC Reference (TestCase-level Scope Drift Guard).
    """
    print("\n[8/8] Testing New QA Capabilities (Assertion-Anchor, Duplicate, Scenario Lint, Live Config, Domain Routing, Phantom AC)...")

    # 1. Assertion-Anchor check (4c): Expected Result đủ dài nhưng KHÔNG có tiêu chí định lượng nào
    vague_tc = TestCase(expected_result="Hệ thống xử lý xong và hiển thị thông báo cho người dùng biết kết quả")
    anchor_issues = lint_test_case(vague_tc)
    assert any(i.issue_type == "Non-Deterministic Expected Result" and "định lượng" in i.description for i in anchor_issues)
    print("  -> Assertion-Anchor check correctly rejected Expected Result without HTTP/status/schema/số liệu!")

    # 2. Duplicate Test Case check (4d): 2 test case trùng tiêu đề + test data
    dup_a = TestCase(testcase_id="TC 01", title="Đăng nhập thành công", test_data='{"user": "a"}')
    dup_b = TestCase(testcase_id="TC 02", title="Đăng nhập thành công", test_data='{"user": "a"}')
    dup_analysis = RequirementAnalysis(acceptance_criteria=[AcceptanceCriterion(ac_id="AC-01")])
    dup_issues = lint_test_suite_coverage(dup_analysis, [dup_a, dup_b])
    assert any(i.issue_type == "Duplicate Test Case" for i in dup_issues)
    print("  -> Duplicate Test Case check correctly flagged 2 identical test cases!")

    # 3. Scenario-Level Lint (Step 7): phantom AC, thiếu kỹ thuật bắt buộc, rò rỉ tên kỹ thuật, thiếu RBT
    scenario_analysis = RequirementAnalysis(
        acceptance_criteria=[AcceptanceCriterion(ac_id="AC-01")],
        product_risks=[ProductRisk(risk_id="RSK-01", risk_level="Critical")]
    )
    leaky_scenario = TestScenario(scenario_id="SC_01", trace_ac_id="AC-99", scenario_title="Kiểm tra BVA giá trị biên", testing_technique="Equivalence Partitioning")
    scenario_issues = lint_scenarios(scenario_analysis, [leaky_scenario])
    scenario_issue_types = [i.issue_type for i in scenario_issues]
    assert "Traceability Gap" in scenario_issue_types  # AC-99 không tồn tại + AC-01 chưa được bao phủ
    assert "Technique Under-Coverage" in scenario_issue_types  # Thiếu Boundary Value Analysis
    assert "Format Violation" in scenario_issue_types  # "BVA" rò rỉ vào scenario_title
    assert "RBT Under-Coverage Violation" in scenario_issue_types  # RSK-01 Critical chưa có scenario trace tới
    print(f"  -> Scenario-Level Linter correctly flagged {len(scenario_issues)} scenario issues: {scenario_issue_types}")

    # 4. Live qa_rules (Step 3): config.yaml là nguồn thật duy nhất, không hardcode rải rác trong code
    live_rules = load_qa_rules()
    yaml_qa_rules = load_config().get("qa_rules", {})
    assert live_rules["min_review_score"] == yaml_qa_rules["min_review_score"] == 95
    assert "verify it works" in live_rules["banned_vague_words"]
    print(f"  -> qa_rules đọc trực tiếp từ config.yaml, min_review_score={live_rules['min_review_score']} (không hardcode)!")

    # 5. Domain Pack Routing (Step 1/5): đúng domain pack theo từ khóa, mặc định 'api-platform' khi không khớp
    assert resolve_domain_pack("FinTech & Banking (Napas, VietQR)") == "fintech-banking"
    assert resolve_domain_pack("E-Commerce & Retail (Cart, Checkout)") == "ecommerce-retail"
    assert resolve_domain_pack("Một tính năng nội bộ không rõ ngành") == "api-platform"
    assert "Double-Entry" in load_domain_pack("FinTech & Banking")
    print("  -> Domain Pack routing chọn đúng pack theo từ khóa và fallback 'api-platform' khi không khớp!")

    # 6. Phantom AC Reference at TestCase level (Scope Drift Guard): TC trỏ tới AC không tồn tại trong analysis
    scope_analysis = RequirementAnalysis(acceptance_criteria=[AcceptanceCriterion(ac_id="AC-01")])
    onscope_tc = TestCase(testcase_id="TC 01", title="Rút tiền thành công trước EOD", group_feature="1. Chặn rút tiền EOD (AC-01)", note="Trace: AC-01")
    drifted_tc = TestCase(testcase_id="TC 02", title="Chuyển tiền Napas thành công (AC-99)", group_feature="2. Chuyển tiền Napas (AC-99)", note="Trace: AC-99")
    drift_issues = lint_test_suite_coverage(scope_analysis, [onscope_tc, drifted_tc])
    phantom_tc_issues = [i for i in drift_issues if i.issue_type == "Scope Drift / Phantom AC Reference"]
    assert len(phantom_tc_issues) == 1 and phantom_tc_issues[0].target_tc_id == "TC 02" and "AC-99" in phantom_tc_issues[0].description
    assert not any(i.issue_type == "Scope Drift / Phantom AC Reference" for i in lint_test_suite_coverage(scope_analysis, [onscope_tc]))
    print("  -> Phantom AC Reference check correctly flagged Test Case tracing a non-existent AC (Scope Drift)!")


def test_clarification_gate():
    """
    Kiểm chứng cơ chế Hard-Stop Clarification Gate:
    - Deterministic detector + gate (Step 1) bắt buộc bật needs_user_clarification khi thiếu API sample / message.
      Mặc định assume User Story là cho API; nếu tài liệu có yếu tố UI mà không tự nhắc tới API thì bỏ qua nhóm câu hỏi API.
      Khi là câu chuyện API: bắt buộc rõ CẢ request lẫn response. Message: bắt buộc rõ CẢ luồng thành công lẫn thất bại.
    - Waiver phrases ("KHÔNG CÓ API" / "KHÔNG CÓ MESSAGE") miễn trừ nhóm câu hỏi tương ứng.
    - Fabricated-message linter check (Step 4) chỉ flag Critical khi message KHÔNG có căn cứ trong tài liệu gốc.
    - Generator marker invariant (Step 3): test case đánh dấu PENDING CLARIFICATION nhưng không kèm câu hỏi
      thì generate_test_cases() phải tự tổng hợp 1 câu hỏi thay thế.
    """
    print("\n[9/9] Testing Clarification Gate (Missing API Sample / Message)...")
    from unittest.mock import patch
    from src.core.clarification import (
        detect_missing_artifacts, apply_clarification_gate,
        MISSING_API_REQUEST_QUESTION, MISSING_API_RESPONSE_QUESTION,
        MISSING_SUCCESS_MESSAGE_QUESTION, MISSING_ERROR_MESSAGE_QUESTION,
    )
    from src.agents.testcase_generator import generate_test_cases, BatchTestSuiteResponse

    # 1. Mặc định: concept của User Story được ASSUME LÀ CHO API -> tài liệu không nhắc UI lẫn API
    #    vẫn bị hỏi đủ 4 điểm: request, response, message thành công, message lỗi.
    bare = "Chặn rút tiền trong giờ EOD 18h"
    assert detect_missing_artifacts(bare) == [
        MISSING_API_REQUEST_QUESTION, MISSING_API_RESPONSE_QUESTION,
        MISSING_SUCCESS_MESSAGE_QUESTION, MISSING_ERROR_MESSAGE_QUESTION,
    ], detect_missing_artifacts(bare)

    # 1b. Tài liệu có yếu tố UI nhưng KHÔNG nhắc tới API -> bỏ qua nhóm câu hỏi API, vẫn hỏi đủ 2 message
    ui_only = bare + " Yêu cầu hiển thị thông tin trên màn hình ứng dụng."
    assert detect_missing_artifacts(ui_only) == [MISSING_SUCCESS_MESSAGE_QUESTION, MISSING_ERROR_MESSAGE_QUESTION], detect_missing_artifacts(ui_only)

    # 1c. Tài liệu UI nhưng CÓ nhắc tới việc gọi API -> hỏi lại nhóm API (2) cộng nhóm message (2)
    ui_with_api = ui_only + " Màn hình gọi API để xử lý giao dịch."
    assert len(detect_missing_artifacts(ui_with_api)) == 4, detect_missing_artifacts(ui_with_api)

    # 1d. Có request nhưng CHƯA có response -> chỉ còn thiếu response (request đã thỏa)
    request_only = ui_with_api + ' Gọi POST /api/v1/withdraw body {"amount": 1000}.'
    assert detect_missing_artifacts(request_only) == [MISSING_API_RESPONSE_QUESTION, MISSING_SUCCESS_MESSAGE_QUESTION, MISSING_ERROR_MESSAGE_QUESTION]

    # 1e. Đủ CẢ request, response, message thành công lẫn message lỗi -> không còn câu hỏi nào
    full_api = (
        request_only + ' Response trả về HTTP 200 kèm thông báo "Rút tiền thành công". '
        'Trường hợp lỗi trả về HTTP 400 mã lỗi CV_043 kèm thông báo "Giao dịch thất bại do tài khoản bị khóa".'
    )
    assert detect_missing_artifacts(full_api) == [], detect_missing_artifacts(full_api)
    assert detect_missing_artifacts(ui_with_api + " KHÔNG CÓ API, KHÔNG CÓ MESSAGE") == []

    # 1f. Waiver TỰ DO (không cần đúng khuôn mẫu "KHÔNG CÓ API/MESSAGE") vẫn được chấp nhận
    free_form_waiver = bare + " Tính năng này hiện chưa có API nào cả, cũng chưa quy định message riêng gì hết."
    assert detect_missing_artifacts(free_form_waiver) == [], detect_missing_artifacts(free_form_waiver)
    qna_style_waiver = bare + " Message: N/A. API: not applicable."
    assert detect_missing_artifacts(qna_style_waiver) == [], detect_missing_artifacts(qna_style_waiver)

    gated_default = apply_clarification_gate(RequirementAnalysis(feature_name="X"), bare)
    assert gated_default.needs_user_clarification is True and len(gated_default.clarification_questions) == 4
    gated_ui = apply_clarification_gate(RequirementAnalysis(feature_name="Y"), ui_only)
    assert gated_ui.needs_user_clarification is True and gated_ui.clarification_questions == [MISSING_SUCCESS_MESSAGE_QUESTION, MISSING_ERROR_MESSAGE_QUESTION]
    print("  -> Deterministic detector requires BOTH request+response and BOTH success+error message; UI-only docs skip the API group; waivers honoured!")

    # 2. Fabricated-message linter check: message bịa đặt bị Critical-flag, message có căn cứ thì sạch
    msg = "Giao dịch của bạn đã bị từ chối do hệ thống đang khóa"
    fab_tc = TestCase(testcase_id="TC 01", title="Rút tiền bị chặn (AC-01)", note="Trace: AC-01",
                       expected_result='HTTP 400, message "' + msg + '"')
    fab_analysis = RequirementAnalysis(acceptance_criteria=[AcceptanceCriterion(ac_id="AC-01", title="Chặn rút tiền EOD")])
    bad = [i for i in lint_test_suite_coverage(fab_analysis, [fab_tc], raw_content="Chặn rút tiền trong giờ EOD, trả HTTP 400.")
           if i.issue_type == "Fabricated Message / Ungrounded Value"]
    assert len(bad) == 1 and bad[0].severity == "Critical" and bad[0].target_tc_id == "TC 01", bad
    good = [i for i in lint_test_suite_coverage(fab_analysis, [fab_tc], raw_content='Khi bị chặn, hiển thị message "' + msg + '".')
            if i.issue_type == "Fabricated Message / Ungrounded Value"]
    assert good == [], good
    assert [i for i in lint_test_suite_coverage(fab_analysis, [fab_tc]) if i.issue_type == "Fabricated Message / Ungrounded Value"] == []
    print("  -> Fabricated-message linter check flags invented messages Critical and clears grounded ones!")

    # 3. Generator marker invariant: PENDING CLARIFICATION marker without a question synthesizes one
    gen_analysis = RequirementAnalysis(feature_name="Chặn rút tiền EOD", acceptance_criteria=[AcceptanceCriterion(ac_id="AC-01")])
    gen_scenarios = [TestScenario(scenario_id="SC_01", trace_ac_id="AC-01", group_feature="1. Chặn EOD", group_functional="1.1. Luồng chính")]
    marked_tc = TestCase(testcase_id="TC 01", title="Rút tiền bị chặn", expected_result="HTTP 400", note="AC-01 | PENDING CLARIFICATION")
    with patch("src.agents.testcase_generator.invoke_structured_llm") as mock_gen:
        mock_gen.return_value = BatchTestSuiteResponse(test_cases=[marked_tc], clarification_questions=[])
        gen_result = generate_test_cases(gen_analysis, gen_scenarios, provider="gemini")
    assert len(gen_result.clarification_questions) == 1, gen_result.clarification_questions
    print("  -> Generator marker invariant: PENDING CLARIFICATION without a question synthesizes one!")


def test_gate_status_reporting():
    """
    Kiểm chứng bug đã sửa: lý do CHƯA ĐẠT Quality Gate hiển thị cho User (CLI/Slack) phải khớp ĐÚNG
    điều kiện gate thật (score >= min_review_score VÀ không Critical/Major), KHÔNG được hardcode
    "Score X/100 < 95" trong khi X đã >= ngưỡng cấu hình và lý do thật là còn issue Critical/Major.
    """
    print("\n[10/10] Testing QA Gate Status Reporting (Score vs. Critical/Major reasons)...")
    from src.agents.reviewer import gate_failure_reasons

    min_score = load_qa_rules()["min_review_score"]

    # 1. Score đạt ngưỡng nhưng còn issue Critical -> lý do phải là Critical, TUYỆT ĐỐI KHÔNG được nói "Score < ngưỡng"
    high_score_critical = ReviewResult(
        passed=False, score=min_score + 1,
        issues=[ReviewIssue(target_tc_id="TC 01", issue_type="X", severity="Critical", description="d")]
    )
    reasons = gate_failure_reasons(high_score_critical)
    assert any("Critical" in r for r in reasons), reasons
    assert not any("chưa đạt ngưỡng" in r for r in reasons), reasons

    # 2. Score dưới ngưỡng, không issue nặng -> lý do phải là Score, đúng số min_score cấu hình (không hardcode 95)
    low_score_clean = ReviewResult(passed=False, score=min_score - 5, issues=[])
    reasons2 = gate_failure_reasons(low_score_clean)
    assert reasons2 == [f"Score {min_score - 5}/100 chưa đạt ngưỡng {min_score}/100"], reasons2

    # 3. Đạt cả điểm lẫn không issue nặng -> không còn lý do nào
    assert gate_failure_reasons(ReviewResult(passed=True, score=min_score, issues=[])) == []
    print(f"  -> gate_failure_reasons() correctly attributes FAILED to Critical/Major issues (not a false score threshold), using live min_review_score={min_score}!")


def test_slack_thread_context_memory():
    """
    Kiểm chứng bug đã sửa: Agent hỏi lại User trong Slack thread (thiếu API sample/message) rồi
    User trả lời trong thread -> Agent PHẢI nhớ ticket/tài liệu gốc và gộp câu trả lời vào đó,
    KHÔNG được coi câu trả lời (thường rất ngắn, ví dụ "KHÔNG CÓ API") như 1 yêu cầu hoàn toàn mới
    rời rạc rồi bị Guardrail từ chối vì "quá ngắn / thiếu tiêu chí nghiệp vụ".
    """
    print("\n[11/11] Testing Slack Thread-Context Memory (Clarification Reply Must Not Forget Original Ticket)...")
    from unittest.mock import patch
    import src.integrations.slack_bot as slack_bot
    from src.agents.testcase_generator import TestCaseGenerationResult

    slack_bot._pending_thread_context.clear()

    class FakeClient:
        def chat_postMessage(self, **kwargs):
            return {"ts": "999.999"}
        def chat_update(self, **kwargs):
            return {"ts": kwargs.get("ts")}
        def files_upload_v2(self, **kwargs):
            pass

    client = FakeClient()
    channel_id = "C123"
    thread_ts = "111.111"
    context_key = f"{channel_id}:{thread_ts}"
    original_ticket_text = (
        "Yeu cau nghiep vu Chuyen tien nhanh Napas hai bon bay: khach hang ca nhan chuyen tien den "
        "so tai khoan ngan hang khac. AC1 so tien toi thieu 10000 toi da 499999999 VND moi giao dich."
    )

    analyze_calls = []
    def fake_analyze_requirements(raw_content, **kwargs):
        analyze_calls.append(raw_content)
        if len(analyze_calls) == 1:
            return RequirementAnalysis(
                feature_name="Chuyen tien Napas", needs_user_clarification=True,
                clarification_questions=["Vui lòng cung cấp API sample (request/response) cho tính năng này."]
            )
        assert "Napas" in raw_content, "Original ticket content was lost on the clarification reply round!"
        assert "KHÔNG CÓ API" in raw_content, "User's clarification reply was lost on the reply round!"
        return RequirementAnalysis(feature_name="Chuyen tien Napas", needs_user_clarification=False)

    with patch("src.integrations.slack_bot.analyze_requirements", side_effect=fake_analyze_requirements), \
         patch("src.integrations.slack_bot.design_test_scenarios", return_value=[TestScenario(scenario_id="SC-01", scenario_title="x", technique="EP")]), \
         patch("src.integrations.slack_bot.generate_test_cases") as mock_gen, \
         patch("src.integrations.slack_bot.review_and_lint_test_suite", return_value=ReviewResult(passed=True, score=100)), \
         patch("src.integrations.slack_bot.export_test_cases_to_excel", return_value=None):
        mock_gen.return_value = TestCaseGenerationResult(test_cases=[TestCase(testcase_id="TC 01", title="x")], clarification_questions=[])

        # Round 1: user tags bot with the original ticket -> Agent must stop and ask, and remember it.
        slack_bot.run_workflow_in_background(client, channel_id, thread_ts, [original_ticket_text], context_key)
        remembered = slack_bot._get_pending_context(context_key)
        assert remembered is not None, "Context was not remembered after asking for clarification!"
        assert "Napas" in remembered

        # Round 2: user replies in-thread with ONLY "KHÔNG CÓ API" (14 chars - fails guardrail standalone).
        slack_bot.run_workflow_in_background(client, channel_id, thread_ts, [remembered, "KHÔNG CÓ API"], context_key)

    assert len(analyze_calls) == 2, f"Expected exactly 2 analysis calls (original + merged reply), got {len(analyze_calls)}"
    assert slack_bot._get_pending_context(context_key) is None, "Context must be cleared once the workflow fully resolves!"
    print("  -> Clarification reply correctly merged with original ticket; thread context cleared once resolved!")



if __name__ == "__main__":
    test_file_parser()
    test_linter()
    test_excel_exporter()
    test_graph_compilation()
    test_agent_invocations()
    test_multi_domain_support()
    test_linter_dead_checks_regression()
    test_new_qa_capabilities()
    test_clarification_gate()
    test_gate_status_reporting()
    test_slack_thread_context_memory()
    print("\n✅ All component tests PASSED!")
