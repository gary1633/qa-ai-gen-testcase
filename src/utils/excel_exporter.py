import os
import re
from datetime import datetime
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image
from src.core.models import RequirementAnalysis, TestCase
from src.core.clarification import PENDING_CLARIFICATION_MARKER

# Style definitions matching EF_TestCases.xlsx
FONT_FAMILY = "Be Vietnam Pro"
FONT_CODE = "Consolas"

HEADER_FONT = Font(name=FONT_FAMILY, size=10, bold=True)
DATA_FONT = Font(name=FONT_FAMILY, size=10, bold=False)
ID_FONT = Font(name=FONT_FAMILY, size=10, bold=True)
CODE_FONT = Font(name=FONT_CODE, size=9, bold=False)

GROUP_L1_FONT = Font(name=FONT_FAMILY, size=11, bold=True, color="000000")
GROUP_L1_FILL = PatternFill(start_color="FFD5A6BD", end_color="FFD5A6BD", fill_type="solid")

GROUP_L2_FONT = Font(name=FONT_FAMILY, size=10, bold=True, color="000000")
GROUP_L2_FILL = PatternFill(start_color="FFEAD1DC", end_color="FFEAD1DC", fill_type="solid")

PENDING_FILL = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style='thin', color='FFD9D9D9'),
    right=Side(style='thin', color='FFD9D9D9'),
    top=Side(style='thin', color='FFD9D9D9'),
    bottom=Side(style='thin', color='FFD9D9D9')
)

# Danh sách các Sheet dùng chung của dự án bắt buộc giữ lại
STANDARD_COMMON_SHEETS = [
    'Workstream Progress Summary',
    'Tài liệu tổng hợp',
    'Template phiếu KTNV',
    'Daily Execution Log',
    'Checklist Report',
    'RAW JIRA Bug'
]

LOGO_PATH = os.path.join(os.path.dirname(__file__), "../../assets/logo.png")

def format_cell_json_text(text: str) -> str:
    """Tự động phát hiện và format đẹp các đoạn JSON trong Steps, Expected Result, Test Data."""
    if not text or not isinstance(text, str):
        return text or ""
    
    text = text.strip()
    
    # 1. Nếu toàn bộ ô là JSON
    if (text.startswith("{") and text.endswith("}")) or (text.startswith("[") and text.endswith("]")):
        try:
            import json
            parsed = json.loads(text)
            return json.dumps(parsed, indent=2, ensure_ascii=False)
        except Exception:
            pass
            
    # 2. Nếu là text có chứa đoạn JSON lồng bên trong (như hình mẫu)
    def replacer(match):
        json_str = match.group(0)
        try:
            import json
            parsed = json.loads(json_str)
            return json.dumps(parsed, indent=2, ensure_ascii=False)
        except Exception:
            return json_str

    pattern = r'(\{(?:[^{}]|(?:\{[^{}]*\}))*\})'
    formatted_text = re.sub(pattern, replacer, text)
    return formatted_text


def sanitize_filename(name: str) -> str:
    """Tạo tên file an toàn từ tên tính năng."""
    clean = re.sub(r'[\\/*?:\[\]"<>|]', '_', name)
    clean = clean.replace(" ", "_")
    clean = re.sub(r'_+', '_', clean).strip('_')
    return clean[:50] if clean else "TestSuite"


def sanitize_sheet_name(name: str) -> str:
    """Loại bỏ ký tự không hợp lệ trong Excel sheet name và giới hạn 30 ký tự."""
    clean = re.sub(r'[\\/*?:\[\]]', '_', name)
    clean = clean.strip()
    return clean[:30] if clean else "QA_TestSuite"


def get_default_template_path() -> str:
    """Ưu tiên dùng EF_TestCases.xlsx (bản chuẩn có logo), fallback sang Template Testsuite.xlsx."""
    if os.path.exists("EF_TestCases.xlsx"):
        return "EF_TestCases.xlsx"
    elif os.path.exists("Template Testsuite.xlsx"):
        return "Template Testsuite.xlsx"
    return "EF_TestCases.xlsx"


def export_test_cases_to_excel(
    analysis: RequirementAnalysis,
    test_cases: List[TestCase],
    template_path: Optional[str] = None,
    output_path: Optional[str] = None,
    target_sheet_name: Optional[str] = None,
    *,
    pending_clarifications: Optional[List[str]] = None
) -> str:
    """
    Tạo một file Excel MỚI RIÊNG BIỆT (mặc định trong thư mục outputs/):
    - Sử dụng EF_TestCases.xlsx làm template chuẩn có logo và biểu đồ
    - GIỮ LẠI tất cả các sheet báo cáo chung (Workstream Progress Summary, Tài liệu tổng hợp, Daily Execution Log...)
    - XÓA các sheet test case demo cũ (EF_BLOCKADE_SA, EF_Calculate_Savings_Interest, fisa_blockade_options...)
    - TẠO sheet test case mới cho tính năng hiện tại và đặt ngay sau sheet Progress Summary
    - TỰ ĐỘNG CẬP NHẬT công thức liên kết trong sheet Workstream Progress Summary
    - NHÚNG LOGO công ty vào sheet Progress Summary
    """
    actual_template = template_path or get_default_template_path()
    if not os.path.exists(actual_template):
        raise FileNotFoundError(f"Không tìm thấy file template tại: {actual_template}")

    # 1. Xác định đường dẫn file đích
    clean_feature_name = sanitize_filename(analysis.feature_name)
    if output_path:
        dest_path = output_path
    else:
        output_dir = "outputs"
        os.makedirs(output_dir, exist_ok=True)
        dest_path = os.path.join(output_dir, f"Testsuite_{clean_feature_name}.xlsx")

    parent_dir = os.path.dirname(dest_path)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)

    # 2. Đọc file Template gốc
    wb = openpyxl.load_workbook(actual_template)
    
    # Tìm sheet template làm mẫu sao chép
    base_sheet_name = None
    candidate_base_sheets = ['Template phiếu KTNV', 'EF_BLOCKADE_SA', 'fisa_blockade_options']
    for candidate in candidate_base_sheets:
        if candidate in wb.sheetnames:
            base_sheet_name = candidate
            break
            
    if not base_sheet_name:
        # Lấy sheet thứ 3 hoặc sheet bất kỳ có cấu trúc testcase
        base_sheet_name = wb.sheetnames[-1]

    # 3. Đặt tên sheet mới & Copy từ template sheet
    desired_name = target_sheet_name or sanitize_sheet_name(analysis.feature_name)
    final_sheet_name = desired_name
    counter = 1
    while final_sheet_name in wb.sheetnames:
        final_sheet_name = f"{desired_name[:26]}_{counter}"
        counter += 1

    base_ws = wb[base_sheet_name]
    ws = wb.copy_worksheet(base_ws)
    ws.title = final_sheet_name

    # 4. Xóa các sheet demo cũ không nằm trong danh sách dùng chung
    for sname in list(wb.sheetnames):
        if sname not in STANDARD_COMMON_SHEETS and sname != final_sheet_name:
            del wb[sname]

    # 5. Di chuyển sheet mới lên vị trí thứ 2 (ngay sau Workstream Progress Summary)
    if 'Workstream Progress Summary' in wb.sheetnames:
        sum_idx = wb.sheetnames.index('Workstream Progress Summary')
        cur_idx = wb.sheetnames.index(final_sheet_name)
        wb.move_sheet(ws, offset=(sum_idx + 1) - cur_idx)

    # 6. Cập nhật Metadata cho Sheet tính năng mới (Rows 10-14, Col C)
    today_str = datetime.now().strftime("%d/%m/%Y")
    ws.cell(10, 3).value = analysis.app_name
    ws.cell(11, 3).value = analysis.version
    ws.cell(12, 3).value = analysis.feature_name
    ws.cell(13, 3).value = analysis.jira_or_doc_link or "N/A"
    ws.cell(14, 3).value = today_str

    # 7. Xóa placeholder từ dòng 22 trở xuống
    max_init_row = ws.max_row
    if max_init_row >= 22:
        ws.delete_rows(22, max_init_row - 21 + 10)

    # 8. Đổ dữ liệu Test Case có gom nhóm L1 / L2
    current_row = 22
    current_group_feature = None
    current_group_functional = None
    first_tc_row = None  # Neo dòng đầu tiên của bộ Test Case, dùng cho công thức tự đánh số cột A

    for tc in test_cases:
        # Gom nhóm L1: Group Feature
        if tc.group_feature and tc.group_feature != current_group_feature:
            current_group_feature = tc.group_feature
            current_group_functional = None
            
            for col_idx in range(1, 15):
                cell = ws.cell(current_row, col_idx)
                cell.fill = GROUP_L1_FILL
                cell.border = THIN_BORDER
                if col_idx == 2:
                    cell.value = current_group_feature
                    cell.font = GROUP_L1_FONT
                    cell.alignment = Alignment(vertical="center", horizontal="left")
            current_row += 1

        # Gom nhóm L2: Group Functional
        if tc.group_functional and tc.group_functional != current_group_functional:
            current_group_functional = tc.group_functional
            
            for col_idx in range(1, 15):
                cell = ws.cell(current_row, col_idx)
                cell.fill = GROUP_L2_FILL
                cell.border = THIN_BORDER
                if col_idx == 2:
                    cell.value = current_group_functional
                    cell.font = GROUP_L2_FONT
                    cell.alignment = Alignment(vertical="center", horizontal="left")
            current_row += 1

        # Ghi 14 cột của Test Case (Format đẹp các khối JSON)
        if first_tc_row is None:
            first_tc_row = current_row
        row_values = [
            f'=IF(B{current_row}<>"", "TC " & TEXT(ROW()-ROW($B${first_tc_row})+1, "00"), "")',
            tc.title,
            tc.preconditions,
            format_cell_json_text(tc.steps),
            format_cell_json_text(tc.expected_result),
            tc.actual_result or "",
            format_cell_json_text(tc.test_data),
            tc.creator or "QA Agent (RBT)",
            tc.test_date or today_str,
            tc.test_status or "Not Test",
            tc.priority or "High",
            tc.plan_execution or "Sprint Release",
            tc.executed_date or "",
            tc.note or ""
        ]

        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(current_row, col_idx)
            cell.value = val
            cell.border = THIN_BORDER

            if col_idx == 1:
                cell.font = ID_FONT
                cell.alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)
            elif col_idx in [2, 3, 4]:
                cell.font = DATA_FONT
                cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
            elif col_idx == 5:
                cell.font = CODE_FONT if "{" in str(val) else DATA_FONT
                cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
            elif col_idx in [8, 9, 10, 11, 12, 13]:
                cell.font = DATA_FONT
                cell.alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)
            else:
                cell.font = DATA_FONT
                cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
        if PENDING_CLARIFICATION_MARKER in (tc.note or ""):
            for col_idx in range(1, 15):
                ws.cell(current_row, col_idx).fill = PENDING_FILL

        current_row += 1

    end_data_row = max(current_row - 1, 22)

    # 9. Cập nhật lại công thức dòng 19 trong sheet testcase
    ws.cell(19, 1).value = f'=COUNTIF(A22:A{end_data_row},"TC*")'
    ws.cell(19, 2).value = f'=COUNTIF(J22:J{end_data_row},"Passed")'
    ws.cell(19, 3).value = f'=COUNTIF(J22:J{end_data_row},"Failed")'
    ws.cell(19, 4).value = f'=COUNTIF(J22:J{end_data_row},"Blocked")'
    ws.cell(19, 5).value = f'=COUNTIF(J22:J{end_data_row},"Not Test")'
    ws.cell(19, 6).value = f'=A19-SUM(B19:E19)'

    # 10. Cập nhật Sheet 'Workstream Progress Summary'
    if 'Workstream Progress Summary' in wb.sheetnames:
        ws_sum = wb['Workstream Progress Summary']
        ws_sum.cell(8, 4).value = analysis.app_name
        ws_sum.cell(9, 4).value = analysis.version
        ws_sum.cell(10, 4).value = today_str
        
        # Cập nhật Row 15 liên kết sang sheet tính năng mới
        ws_sum.cell(15, 2).value = 1
        ws_sum.cell(15, 3).value = final_sheet_name
        ws_sum.cell(15, 4).value = f"='{final_sheet_name}'!A19"
        ws_sum.cell(15, 5).value = f"='{final_sheet_name}'!B19"
        ws_sum.cell(15, 6).value = f"='{final_sheet_name}'!C19"
        ws_sum.cell(15, 7).value = f"='{final_sheet_name}'!D19"
        ws_sum.cell(15, 8).value = f"='{final_sheet_name}'!E19"
        ws_sum.cell(15, 9).value = "=SUM(E15:H15)-G15"
        ws_sum.cell(15, 10).value = "=D15-I15"
        ws_sum.cell(15, 11).value = '=IFERROR(I15/D15,"")'
        
        # Xóa dữ liệu các dòng demo cũ (Row 16, 17)
        for r_clear in range(16, 18):
            for c_clear in range(2, 12):
                ws_sum.cell(r_clear, c_clear).value = None

        # Cập nhật công thức Tổng dòng 18
        ws_sum.cell(18, 4).value = "=SUM(D15:D15)"
        ws_sum.cell(18, 5).value = "=D18-F18"
        ws_sum.cell(18, 6).value = "=SUM(F15:F15)"
        ws_sum.cell(18, 7).value = "=SUM(G15:G15)"
        ws_sum.cell(18, 8).value = "=SUM(H15:H15)"
        ws_sum.cell(18, 9).value = "=SUM(I15:I15)"
        ws_sum.cell(18, 10).value = "=SUM(J15:J15)"
        ws_sum.cell(18, 11).value = '=IFERROR(I18/D18,"")'

        # Logo và biểu đồ đã có sẵn nguyên bản từ template EF_TestCases.xlsx

    if pending_clarifications:
        q_ws = wb.create_sheet(title="Cần làm rõ (Pending)")
        q_ws.cell(1, 1).value = f"CÂU HỎI CẦN USER / PO / BA LÀM RÕ - {analysis.feature_name}"
        q_ws.cell(1, 1).font = GROUP_L1_FONT
        q_ws.cell(2, 1).value = (
            "Các test case được tô màu vàng và có ghi chú PENDING CLARIFICATION trong sheet test case "
            "đang thiếu API sample / message chính xác. KHÔNG dùng làm bản chính thức trước khi chốt các câu hỏi dưới đây."
        )
        for i, q in enumerate(pending_clarifications, start=1):
            cell = q_ws.cell(3 + i, 1)
            cell.value = f"{i}. {q}"
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
        q_ws.column_dimensions["A"].width = 140

    # Đặt sheet test case mới làm Active Sheet
    wb.active = ws

    # Lưu ra file mới
    wb.save(dest_path)
    return dest_path
